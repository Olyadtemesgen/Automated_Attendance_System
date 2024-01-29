import os
import zipfile
import shutil
import openpyxl
import face_recognition
import matplotlib.pyplot as plt
import numpy as np
from keras.preprocessing import image
from PIL import Image
from openpyxl.styles import Alignment

def add_student(zipFileName: str):
    student_name = os.path.splitext(zipFileName)[0]
    tobesplited = os.path.splitext(zipFileName)[0].split('_')[-1]
    name, section, gender, id = os.path.splitext(tobesplited)[0].split('-')

    name = name.split('/')[-1]
    add_student_to_excel("/home/olitye/Code/AI/CNN/attendance/Students_Attendance.xlsx", name, id, section, gender)

    with zipfile.ZipFile(zipFileName, 'r') as zip_ref:
        zip_ref.extractall(student_name)


    extracted_folder = os.path.join(student_name, student_name)
    destination_directory = '/home/olitye/Code/AI/CNN/Images/Training-Images'

    # Move the contents of the extracted folder to the destination directory
    folder_name = ""
    extractedName = student_name.split('/')[-1]

    shutil.move(extracted_folder, destination_directory)
    # rename the folder in the destination to tobessplited
    os.rename(destination_directory + '/' + extractedName, destination_directory + '/' + tobesplited)
 

    #remove the zip file 
    os.remove(zipFileName)
    # Go Through the images and detect only one faces and save it to that the same image file
    # Get the path to the extracted folder (student_name/student_name)

    # split the student name to get the name only
    student_name = student_name.split('/')[-1]
    print("student_nme", student_name)

    images_length = len(os.listdir(destination_directory + '/' + tobesplited))

    if not os.path.exists('/home/olitye/Code/AI/CNN/Images/Testing-Images/' + tobesplited):
        os.makedirs('/home/olitye/Code/AI/CNN/Images/Testing-Images/' + tobesplited)
    
    testing_destination = '/home/olitye/Code/AI/CNN/Images/Testing-Images/' + tobesplited

    for image_path in os.listdir(destination_directory + '/' + tobesplited)[:images_length]:

        # move the folder out. 
        # shutil.move(destination_directory + '/' + student_name + '/' + image_path, destination_directory)

        images_length -= 1
        images = face_recognition.load_image_file(destination_directory + '/' + tobesplited + '/' + image_path)

        face_locations = face_recognition.face_locations(images)
        # recognize_faces(face_locations, images)
        print("I found {} face(s) in this photograph.".format(len(face_locations)))
        images_name = []

        if len(face_locations) == 0:
            images_length -= 1
            continue

        for idx, face_location in enumerate(face_locations):
            top, right, bottom, left = face_location

            top = max(0, top - 300)
            left = max(0, left - 300)
         
            right = min(images.shape[1], right + 300)
            bottom = min(images.shape[0], bottom + 300)

            face_image = images[top:bottom, left:right]

            pil_image = Image.fromarray(face_image)

            if images_length > 1:
                pil_image.save(destination_directory + '/' + tobesplited + '/' + image_path)
            
            else:
            
                pil_image.save(testing_destination + '/' + image_path)
                os.remove(destination_directory + '/' + tobesplited + '/' + image_path)
            break
           


def is_student_unique(sheet, student_name, student_id, section, gender):

    for row in sheet.iter_rows(min_row=2):
        existing_name = row[0].value
        existing_id = row[1].value
        existing_section = row[2].value
        existing_gender = row[3].value

        # Check if the student already exists in the sheet
        if (
            existing_name == student_name
            and existing_id == student_id
            and existing_section == section
            and existing_gender == gender
        ):
            return False

    return True

def add_student_to_excel(excel_file_path, student_name, student_id, section, gender):
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path)

    # Select the active sheet (you may need to modify this based on your Excel file structure)
    sheet = wb.active

    # Check if the student is already in the Excel file
    if not is_student_unique(sheet, student_name, student_id, section, gender):
        print("Student already exists in the Excel file.")
        wb.close()
        return

    # Find the column index of the first empty row
    row_index = sheet.max_row + 1

    # Write the student information in separate columns
    sheet.cell(row=row_index, column=1).value = student_name
    sheet.cell(row=row_index, column=2).value = student_id
    sheet.cell(row=row_index, column=3).value = section
    sheet.cell(row=row_index, column=4).value = gender

    # sheet.cell(row=unique_row_index, column=name_column_index).value = student_name
    # sheet.cell(row=unique_row_index, column=name_column_index + 1).value = student_id
    # sheet.cell(row=unique_row_index, column=name_column_index + 2).value = section
    # sheet.cell(row=unique_row_index, column=name_column_index + 3).value = gender

    # Set alignment for the new row
    for col in sheet.iter_cols(min_col=1, max_col=1 + 3):
        for cell in col:
            cell.alignment = Alignment(horizontal='center')

    # Save the modified Excel file
    wb.save(excel_file_path)
    wb.close()
