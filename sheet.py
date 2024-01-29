import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def mark_attendance(student_name, student_id, section, attendance_date, gender):
    excel_file_path = '/home/olitye/Code/AI/CNN/attendance/Students_Attendance.xlsx'  # Replace with the actual path to your Excel file

    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path)

    # Select the active sheet (you may need to modify this based on your Excel file structure)
    sheet = wb.active

    # Find the column index of the student name
    name_column_index = None
    for col in sheet.iter_cols():
        if col[0].value == "Name":
            name_column_index = col[0].column
            break

    if name_column_index is None:
        # If the "Name" column is not found, assume it is the first column (column A)
        name_column_index = 1

    # Find the row index of the student name, ID, and section combination
    unique_row_index = None
    for row in sheet.iter_rows(min_row=2):
        if (
            row[name_column_index - 1].value == student_name
            and row[name_column_index].value == student_id
            and row[name_column_index + 1].value == section
        ):
            unique_row_index = row[name_column_index - 1].row
            break

    if unique_row_index is None:
        # If the unique row is not found, create a new row at the end
        unique_row_index = sheet.max_row + 1

        # Write the student name, ID, and section in separate columns
        sheet.cell(row=unique_row_index, column=name_column_index).value = student_name
        sheet.cell(row=unique_row_index, column=name_column_index + 1).value = student_id
        sheet.cell(row=unique_row_index, column=name_column_index + 2).value = section
        sheet.cell(row=unique_row_index, column=name_column_index + 3).value = gender

        # Set alignment for the new row
        for col in sheet.iter_cols(min_col=name_column_index, max_col=name_column_index + 3):
            for cell in col:
                cell.alignment = Alignment(horizontal='center')

    # Find the column index of the attendance date
    date_column_index = None
    for col in sheet.iter_cols():
        if col[0].value and str(col[0].value) == attendance_date:
            date_column_index = col[0].column
            break

    if date_column_index is None:
        # If the attendance date column is not found, create a new column at the end
        date_column_index = sheet.max_column + 1
        date_column_letter = get_column_letter(date_column_index)

        # Write the attendance date at the top of the column
        sheet[date_column_letter + '1'] = attendance_date
        sheet[date_column_letter + '1'].alignment = Alignment(horizontal='center')

    # Write "Present" below the date for the corresponding student
    present_cell = sheet.cell(row=unique_row_index, column=date_column_index)
    present_cell.value = "Present"
    present_cell.alignment = Alignment(horizontal='center')

    # Adjust the column width for name, ID, section, and date columns based on the longest data plus 5 characters
    columns_to_adjust = [
        name_column_index,
        name_column_index + 1,
        name_column_index + 2,
        date_column_index
    ]
    for column_index in columns_to_adjust:
        column_width = max(len(str(cell.value)) for cell in sheet[get_column_letter(column_index)])
        column_width += 5
        sheet.column_dimensions[get_column_letter(column_index)].width = column_width

    # Save the modified Excel file
    wb.save(excel_file_path)
    wb.close()


# def count_genders(date):
#     # a gender counter who come at that date 
#     excel_file_path = '/home/olitye/Code/AI/CNN/attendance/Students_Attendance.xlsx'  # Replace with the actual path to your Excel file

#     # Load the Excel file
#     wb = openpyxl.load_workbook(excel_file_path)

#     # Select the active sheet (you may need to modify this based on your Excel file structure)
#     sheet = wb.active

#     # Find the column index of the attendance date
#     date_column_index = None
#     for col in sheet.iter_cols():
#         if col[0].value and str(col[0].value) == date:
#             date_column_index = col[0].column
#             break

#     if date_column_index is None:
#         return 0, 0

    
#     male = 0
#     female = 0

    # go to the column that starts with the date date and go and find a row that has "present" value in it and on the second index if it says M add to male and not add to female

def count_males_and_females(date):

    filename = '/home/olitye/Code/AI/CNN/attendance/Students_Attendance.xlsx'  # Replace with the actual path to your Excel file

    wb = openpyxl.load_workbook(filename)

    # Select the active sheet (you may need to modify this based on your Excel file structure)
    sheet = wb.active  # Replace with the actual sheet name

    male_count = 0
    female_count = 0

    # Find the column index for the given date
    date_column = None
    for cell in sheet[1]:
        if cell.value == date:
            date_column = cell.column - 1
            break

    if date_column is None:
        return None, None

    # Count male and female students
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[date_column] == "Present":
            if row[3] == "M":
                male_count += 1
            
            elif row[3] == "F":
                female_count += 1

    return male_count, female_count